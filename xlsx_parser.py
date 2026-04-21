"""Парсер xlsx уведомления о выкупе Wildberries."""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook


@dataclass
class RedemptionItem:
    row_num: int            # № п/п
    article: str            # артикул продавца (КодТов в УПД)
    name: str               # наименование
    quantity: float
    sum_with_vat: float     # сумма выкупа (включая НДС, если облагается)
    vat_rate_src: str       # ставка НДС из уведомления (обычно «без НДС»)
    vat_amount_src: str     # сумма НДС из уведомления
    kiz: str                # КИЗ (код идентификации)


@dataclass
class RedemptionNotice:
    number: str             # «693111100»
    notice_date: date       # дата из заголовка
    items: list[RedemptionItem]


_HEADER_RE = re.compile(
    r"УВЕДОМЛЕНИЕ\s+О\s+ВЫКУПЕ\s*№\s*(\d+)\s+от\s+(\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)


def _float(val) -> float:
    if val is None or val == "—":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_notice_xlsx(src: Union[Path, bytes, BytesIO]) -> RedemptionNotice:
    """Парсит .xlsx из архива уведомления о выкупе."""
    if isinstance(src, (bytes, bytearray)):
        wb = load_workbook(BytesIO(src), data_only=True)
    elif isinstance(src, BytesIO):
        wb = load_workbook(src, data_only=True)
    else:
        wb = load_workbook(str(src), data_only=True)

    ws = wb.active

    # Шапка: «УВЕДОМЛЕНИЕ О ВЫКУПЕ №XXXXX от YYYY-MM-DD» — ищем в первых 10 строках.
    notice_number = ""
    notice_date: Optional[date] = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for v in row:
            if isinstance(v, str):
                m = _HEADER_RE.search(v)
                if m:
                    notice_number = m.group(1)
                    notice_date = datetime.strptime(m.group(2), "%Y-%m-%d").date()
                    break
        if notice_number:
            break

    if not notice_number or not notice_date:
        raise ValueError("Не удалось найти шапку уведомления о выкупе (№ и дата).")

    # Заголовки таблицы — строка 10. Данные — с 11-й.
    items: list[RedemptionItem] = []
    for r in range(11, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        article = ws.cell(r, 2).value
        if article is None:
            continue

        items.append(
            RedemptionItem(
                row_num=int(no) if isinstance(no, (int, float)) else len(items) + 1,
                article=str(article).strip(),
                name=str(ws.cell(r, 3).value or "").strip(),
                quantity=_float(ws.cell(r, 4).value),
                sum_with_vat=_float(ws.cell(r, 5).value),
                vat_rate_src=str(ws.cell(r, 6).value or "").strip(),
                vat_amount_src=str(ws.cell(r, 7).value or "").strip(),
                kiz=str(ws.cell(r, 8).value or "").strip(),
            )
        )

    return RedemptionNotice(number=notice_number, notice_date=notice_date, items=items)


def extract_xlsx_from_zip(zip_bytes: bytes) -> bytes:
    """Достаёт первый xlsx-файл из zip-архива уведомления о выкупе."""
    with zipfile.ZipFile(BytesIO(zip_bytes)) as z:
        for info in z.infolist():
            # Используем orig_filename — он корректно декодируется, если в архиве стоит флаг UTF-8.
            name = info.orig_filename
            if name.lower().endswith(".xlsx"):
                return z.read(info.filename)
    raise ValueError("В архиве нет .xlsx-файла.")
